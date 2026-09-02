# -*- coding: utf-8 -*-
from kivy.lang import Builder
from kivy.core.window import Window
from kivy.utils import get_color_from_hex
from kivy.metrics import dp
from kivy.uix.screenmanager import ScreenManager, Screen
from kivymd.app import MDApp
from kivymd.uix.dialog import MDDialog
from kivymd.uix.button import MDRaisedButton, MDTextButton
from kivymd.uix.textfield import MDTextField
from kivymd.uix.label import MDLabel
from kivymd.uix.datatables import MDDataTable
from kivymd.uix.boxlayout import MDBoxLayout
from kivymd.uix.scrollview import MDScrollView
from kivymd.uix.floatlayout import MDFloatLayout
import sqlite3
import re
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


# ==================== Android 兼容处理 ====================
# 判断是否在 Android 上运行
IS_ANDROID = False
try:
    from android.permissions import request_permissions, Permission
    IS_ANDROID = True
except ImportError:
    pass

# 获取正确的数据库路径
def get_db_path():
    if IS_ANDROID:
        # Android 上使用应用私有目录
        from android.storage import app_storage_path
        return os.path.join(app_storage_path(), "business.db")
    else:
        return "business.db"

# ==================== 全局配置 ====================
ADMIN_PASSWORD = "432"
EXCEL_HEADERS = ["检定项目", "型号规格", "测量范围", "价格"]
DB_NAME = get_db_path()

# 只在非 Android 环境下设置窗口大小
if not IS_ANDROID:
    Window.size = (360, 640)
Window.clearcolor = get_color_from_hex("#f5f5f5")

# ==================== 数据库初始化 ====================
def init_db():
    try:
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        c.execute("""
            CREATE TABLE IF NOT EXISTS capabilities (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_name TEXT NOT NULL,
                model_spec TEXT NOT NULL,
                measure_range TEXT NOT NULL,
                price TEXT NOT NULL
            )
        """)
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        print(f"数据库初始化失败: {e}")
        return False

def parse_price(price_str):
    nums = re.findall(r'\d+\.?\d*', str(price_str))
    if nums:
        return float(nums[0])
    return 0.0

# ==================== 页面定义 ====================
KV = '''
ScreenManager:
    LoginScreen:
    ClientScreen:
    AdminScreen:
    QuoteScreen:

# 密码登录页
<LoginScreen>:
    name: "login"
    MDBoxLayout:
        orientation: "vertical"
        spacing: dp(30)
        padding: dp(30)
        pos_hint: {"center_x":0.5, "center_y":0.5}

        MDLabel:
            text: "仪器检校价格系统"
            font_style: "H5"
            halign: "center"

        MDTextField:
            id: pwd_input
            hint_text: "请输入管理员密码"
            password: True
            size_hint: 1, None
            height: dp(50)

        MDRaisedButton:
            text: "进入管理后台"
            md_bg_color: "#c0392b"
            size_hint: 1, None
            height: dp(50)
            on_press: root.check_pwd()

        MDRaisedButton:
            text: "普通查询入口"
            md_bg_color: "#2980b9"
            size_hint: 1, None
            height: dp(50)
            on_press: app.root.current = "client"

# 普通查询页
<ClientScreen>:
    name: "client"
    MDBoxLayout:
        orientation: "vertical"
        padding: dp(10)
        spacing: dp(10)

        MDBoxLayout:
            size_hint: 1, None
            height: dp(50)
            spacing: dp(10)

            MDTextField:
                id: search_key
                hint_text: "输入检定项目搜索"
                size_hint: 0.7, 1

            MDRaisedButton:
                text: "查询"
                size_hint: 0.3, 1
                on_press: root.search_data()

        MDLabel:
            id: tip_text
            text: ""
            halign: "center"
            theme_text_color: "Error"

        MDScrollView:
            size_hint: 1, 1
            MDBoxLayout:
                id: table_box
                size_hint: 1, None
                height: dp(800)

        MDRaisedButton:
            text: "批量报价"
            md_bg_color: "#8e44ad"
            size_hint: 1, None
            height: dp(45)
            on_press: app.root.current = "quote"

        MDTextButton:
            text: "管理员入口"
            on_press: app.root.current = "login"

# 管理员页面
<AdminScreen>:
    name: "admin"
    MDBoxLayout:
        orientation: "vertical"
        padding: dp(10)
        spacing: dp(8)

        MDTextField:
            id: e1
            hint_text: "检定项目"
            size_hint: 1, None
            height: dp(45)
        MDTextField:
            id: e2
            hint_text: "型号规格"
            size_hint: 1, None
            height: dp(45)
        MDTextField:
            id: e3
            hint_text: "测量范围"
            size_hint: 1, None
            height: dp(45)
        MDTextField:
            id: e4
            hint_text: "价格"
            size_hint: 1, None
            height: dp(45)

        MDLabel:
            id: edit_tip
            text: ""
            theme_text_color: "Warning"

        MDBoxLayout:
            spacing: dp(5)
            size_hint: 1, None
            height: dp(40)

            MDRaisedButton:
                text: "保存"
                md_bg_color: "#27ae60"
                on_press: root.save_data()
            MDRaisedButton:
                text: "修改"
                md_bg_color: "#f39c12"
                on_press: root.edit_data()
            MDRaisedButton:
                text: "删除"
                md_bg_color: "#e74c3c"
                on_press: root.del_data()

        MDBoxLayout:
            spacing: dp(5)
            size_hint: 1, None
            height: dp(40)

            MDRaisedButton:
                text: "导入Excel"
                md_bg_color: "#9b59b6"
                on_press: root.import_excel()
            MDRaisedButton:
                text: "刷新"
                md_bg_color: "#3498db"
                on_press: root.load_list()

        MDScrollView:
            size_hint: 1, 1
            MDBoxLayout:
                id: admin_table_box
                size_hint: 1, None
                height: dp(600)

# 批量报价页面
<QuoteScreen>:
    name: "quote"
    MDBoxLayout:
        orientation: "vertical"
        padding: dp(10)
        spacing: dp(10)

        MDBoxLayout:
            size_hint: 1, None
            height: dp(45)
            spacing: dp(10)
            MDTextField:
                id: q_search
                hint_text: "搜索检定项目"
                size_hint: 0.7,1
            MDRaisedButton:
                text: "搜索"
                size_hint: 0.3,1
                on_press: root.q_search()

        MDScrollView:
            size_hint: 1, 0.4
            MDBoxLayout:
                id: q_table_box
                size_hint:1, None

        MDBoxLayout:
            size_hint:1, None
            height: dp(45)
            spacing: dp(10)
            MDTextField:
                id: qty_input
                hint_text: "数量"
                text: "1"
                size_hint:0.4,1
            MDRaisedButton:
                text: "加入报价单"
                md_bg_color: "#27ae60"
                size_hint:0.6,1
                on_press: root.add_cart()

        MDLabel:
            text: "报价单列表"
            font_style: "Subtitle1"

        MDScrollView:
            size_hint:1,0.4
            MDBoxLayout:
                id: cart_box
                size_hint:1, None

        MDLabel:
            id: total_text
            text: "总金额：0.00 元"
            font_style: "H6"
            theme_text_color: "Error"

        MDBoxLayout:
            spacing: dp(5)
            size_hint:1, None
            height: dp(45)
            MDRaisedButton:
                text: "删除选中"
                md_bg_color: "#e74c3c"
                on_press: root.del_cart()
            MDRaisedButton:
                text: "清空"
                md_bg_color: "#95a5a6"
                on_press: root.clear_cart()

        MDRaisedButton:
            text: "导出Excel报价单"
            md_bg_color: "#8e44ad"
            size_hint:1, None
            height: dp(45)
            on_press: root.export_excel()

        MDTextButton:
            text: "返回查询页"
            on_press: app.root.current = "client"
'''

# ==================== 页面逻辑 ====================
class LoginScreen(Screen):
    def check_pwd(self):
        pwd = self.ids.pwd_input.text.strip()
        if pwd == ADMIN_PASSWORD:
            self.manager.current = "admin"
        else:
            self.show_msg("密码错误！")

    def show_msg(self, txt):
        dialog = MDDialog(text=txt, buttons=[MDRaisedButton(text="确定", on_press=lambda x: dialog.dismiss())])
        dialog.open()

class ClientScreen(Screen):
    def on_enter(self):
        self.search_data()

    def search_data(self):
        key = self.ids.search_key.text.strip()
        self.ids.table_box.clear_widgets()
        try:
            conn = sqlite3.connect(DB_NAME)
            c = conn.cursor()
            if key:
                c.execute("SELECT project_name,model_spec,measure_range,price FROM capabilities WHERE project_name LIKE ?", (f"%{key}%",))
            else:
                c.execute("SELECT project_name,model_spec,measure_range,price FROM capabilities")
            res = c.fetchall()
            conn.close()
            if not res:
                self.ids.tip_text.text = "暂不开展此项业务"
                return
            self.ids.tip_text.text = ""
            table = MDDataTable(
                size_hint=(1, None),
                height=dp(400),
                column_data=[("项目", dp(90)), ("型号", dp(90)), ("范围", dp(90)), ("价格", dp(70))],
                row_data=[(i[0],i[1],i[2],i[3]) for i in res]
            )
            self.ids.table_box.add_widget(table)
        except Exception as e:
            self.ids.tip_text.text = f"数据库错误: {str(e)}"

class AdminScreen(Screen):
    editing_id = None
    def on_enter(self):
        self.load_list()

    def load_list(self):
        self.ids.admin_table_box.clear_widgets()
        try:
            conn = sqlite3.connect(DB_NAME)
            c = conn.cursor()
            c.execute("SELECT id,project_name,model_spec,measure_range,price FROM capabilities")
            res = c.fetchall()
            conn.close()
            table = MDDataTable(
                size_hint=(1, None),
                height=dp(350),
                column_data=[("ID", dp(40)), ("项目", dp(80)), ("型号", dp(90)), ("范围", dp(90)), ("价格", dp(60))],
                row_data=[(str(i[0]),i[1],i[2],i[3],i[4]) for i in res]
            )
            self.ids.admin_table_box.add_widget(table)
            self.table = table
        except Exception as e:
            self.show_msg(f"加载数据失败: {str(e)}")

    def save_data(self):
        p1 = self.ids.e1.text.strip()
        p2 = self.ids.e2.text.strip()
        p3 = self.ids.e3.text.strip()
        p4 = self.ids.e4.text.strip()
        if not all([p1,p2,p3,p4]):
            self.show_msg("所有字段不能为空！")
            return
        try:
            conn = sqlite3.connect(DB_NAME)
            c = conn.cursor()
            if self.editing_id:
                c.execute("UPDATE capabilities SET project_name=?,model_spec=?,measure_range=?,price=? WHERE id=?",(p1,p2,p3,p4,self.editing_id))
                self.editing_id = None
                self.ids.edit_tip.text = ""
            else:
                c.execute("INSERT INTO capabilities(project_name,model_spec,measure_range,price) VALUES (?,?,?,?)",(p1,p2,p3,p4))
            conn.commit()
            conn.close()
            self.clear_input()
            self.load_list()
            self.show_msg("保存成功！")
        except Exception as e:
            self.show_msg(f"保存失败: {str(e)}")

    def edit_data(self):
        if not hasattr(self, 'table') or self.table is None:
            self.show_msg("请先加载数据！")
            return
        row = self.table.current_row
        if not row:
            self.show_msg("请选择一行数据！")
            return
        self.editing_id = int(row[0])
        self.ids.e1.text = row[1]
        self.ids.e2.text = row[2]
        self.ids.e3.text = row[3]
        self.ids.e4.text = row[4]
        self.ids.edit_tip.text = f"正在修改ID:{self.editing_id}"

    def del_data(self):
        if not hasattr(self, 'table') or self.table is None:
            self.show_msg("请先加载数据！")
            return
        row = self.table.current_row
        if not row:
            self.show_msg("请选择一行！")
            return
        try:
            conn = sqlite3.connect(DB_NAME)
            c = conn.cursor()
            c.execute("DELETE FROM capabilities WHERE id=?",(int(row[0]),))
            conn.commit()
            conn.close()
            self.load_list()
            self.show_msg("删除成功！")
        except Exception as e:
            self.show_msg(f"删除失败: {str(e)}")

    def clear_input(self):
        self.ids.e1.text = ""
        self.ids.e2.text = ""
        self.ids.e3.text = ""
        self.ids.e4.text = ""

    def import_excel(self):
        # Android 上 filechooser 可能不支持，提示用户
        if IS_ANDROID:
            self.show_msg("Android 上导入功能暂不支持，请在电脑上操作")
            return
        def select_file(path):
            if not path:
                return
            try:
                wb = load_workbook(path)
                ws = wb.active
                headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                mapping = {}
                for idx, h in enumerate(headers):
                    if h in EXCEL_HEADERS:
                        mapping[h] = idx
                if len(mapping)!=4:
                    self.show_msg("Excel格式错误！")
                    return
                cols = [mapping[h] for h in EXCEL_HEADERS]
                conn = sqlite3.connect(DB_NAME)
                c = conn.cursor()
                cnt = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    vals = [str(row[i]).strip() if row[i] else "" for i in cols]
                    if all(vals):
                        c.execute("INSERT INTO capabilities VALUES (NULL,?,?,?,?)", vals)
                        cnt +=1
                conn.commit()
                conn.close()
                self.load_list()
                self.show_msg(f"成功导入{cnt}条数据")
            except Exception as e:
                self.show_msg(f"导入失败：{str(e)}")
        try:
            filechooser.open_file(on_selection=select_file)
        except Exception as e:
            self.show_msg(f"文件选择器错误: {str(e)}")

    def show_msg(self, txt):
        dialog = MDDialog(text=txt, buttons=[MDRaisedButton(text="确定", on_press=lambda x: dialog.dismiss())])
        dialog.open()

class QuoteScreen(Screen):
    cart_list = []
    def on_enter(self):
        self.q_search()
        self.refresh_cart()

    def q_search(self):
        key = self.ids.q_search.text.strip()
        self.ids.q_table_box.clear_widgets()
        try:
            conn = sqlite3.connect(DB_NAME)
            c = conn.cursor()
            if key:
                c.execute("SELECT project_name,model_spec,measure_range,price FROM capabilities WHERE project_name LIKE ?",(f"%{key}%",))
            else:
                c.execute("SELECT project_name,model_spec,measure_range,price FROM capabilities")
            res = c.fetchall()
            conn.close()
            table = MDDataTable(
                size_hint=(1, None),
                height=dp(220),
                column_data=[("项目",dp(80)),("型号",dp(80)),("范围",dp(80)),("价格",dp(60))],
                row_data=res
            )
            self.ids.q_table_box.add_widget(table)
            self.q_table = table
        except Exception as e:
            self.show_msg(f"查询失败: {str(e)}")

    def add_cart(self):
        if not hasattr(self, 'q_table') or self.q_table is None:
            self.show_msg("请先搜索数据！")
            return
        row = self.q_table.current_row
        if not row:
            self.show_msg("请选择项目！")
            return
        try:
            qty = float(self.ids.qty_input.text.strip())
            if qty <=0:
                raise
        except:
            self.show_msg("数量必须是正数！")
            return
        price_num = parse_price(row[3])
        subtotal = price_num * qty
        self.cart_list.append({
            "name":row[0],"model":row[1],"range":row[2],"price_str":row[3],"price":price_num,"qty":qty,"sub":subtotal
        })
        self.refresh_cart()

    def refresh_cart(self):
        self.ids.cart_box.clear_widgets()
        if not self.cart_list:
            self.ids.total_text.text = "总金额：0.00 元"
            return
        rows = []
        total = 0
        for item in self.cart_list:
            rows.append((item["name"],item["model"],item["range"],item["price_str"],str(item["qty"]),f"{item['sub']:.2f}"))
            total += item["sub"]
        table = MDDataTable(
            size_hint=(1, None),
            height=dp(200),
            column_data=[("项目",70),("型号",70),("范围",70),("单价",50),("数量",50),("小计",60)],
            row_data=rows
        )
        self.ids.cart_box.add_widget(table)
        self.ids.total_text.text = f"总金额：{total:.2f} 元"
        self.cart_table = table

    def del_cart(self):
        if not hasattr(self, 'cart_table') or self.cart_table is None:
            self.show_msg("报价单为空！")
            return
        row = self.cart_table.current_row
        if not row:
            self.show_msg("请选择！")
            return
        idx = self.cart_table.row_data.index(row)
        del self.cart_list[idx]
        self.refresh_cart()

    def clear_cart(self):
        self.cart_list.clear()
        self.refresh_cart()

    def export_excel(self):
        if IS_ANDROID:
            self.show_msg("Android 上导出功能暂不支持，请在电脑上操作")
            return
        if not self.cart_list:
            self.show_msg("报价单为空！")
            return
        def save_file(path):
            if not path:
                return
            wb = Workbook()
            ws = wb.active
            ws.title = "报价单"
            headers = ["检定项目","型号规格","测量范围","单价","数量","小计"]
            ws.append(headers)
            for item in self.cart_list:
                ws.append([item["name"],item["model"],item["range"],item["price_str"],item["qty"],round(item["sub"],2)])
            total = sum(i["sub"] for i in self.cart_list)
            ws.append(["合计","","","","",round(total,2)])
            wb.save(path)
            self.show_msg("导出成功！")
        try:
            filechooser.save_file(on_selection=save_file)
        except Exception as e:
            self.show_msg(f"导出失败: {str(e)}")

    def show_msg(self,txt):
        dialog = MDDialog(text=txt, buttons=[MDRaisedButton(text="确定", on_press=lambda x: dialog.dismiss())])
        dialog.open()

# ==================== 注册页面 ====================
class SM(ScreenManager):
    pass

class PriceApp(MDApp):
    def build(self):
        self.theme_cls.primary_palette = "Blue"
        self.theme_cls.theme_style = "Light"
        # 初始化数据库
        if not init_db():
            print("数据库初始化失败，但应用将继续运行")
        return Builder.load_string(KV)

if __name__ == "__main__":
    PriceApp().run()